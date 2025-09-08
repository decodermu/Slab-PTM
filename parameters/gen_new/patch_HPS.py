import openpyxl
import csv

ref_sort = ["E","D","R","K",
            "I","V","L","T",
            "H","M","G","A",
            "C","S","P","N",
            "Q","F","Y","W"]

def create_workbook(csv_path):
    """
    从 csv_path 读取 target_columns，根据 ref_sort排序后写入 residues sheet
    写入列为 ["one","three","MW","q","type"]，type为0~19
    同时返回按 ref_sort 排序后的 lambdas 列表和 sigmas 列表
    """
    target_columns = ["one", "three", "MW", "q", "lambdas", "sigmas"]

    # read csv
    with open(csv_path, mode="r") as file:
        reader = csv.DictReader(file)
        for col in target_columns:
            if col not in reader.fieldnames:
                raise ValueError(f"CSV 文件缺少列: {col}")
        rows = [row for row in reader]

    #ref_sort
    rows_sorted = sorted(rows, key=lambda r: ref_sort.index(r["one"]))

    # get l and s
    lambdas_list = [float(r["lambdas"]) for r in rows_sorted]
    sigmas_list  = [float(r["sigmas"]) for r in rows_sorted]

    #create workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "residues"

    #write to residues
    sheet.append(["one", "three", "MW", "q", "type"])
    for idx, row in enumerate(rows_sorted):
        sheet.append([row["one"], row["three"], row["MW"], row["q"], idx])

    return workbook, lambdas_list, sigmas_list

def read_factors(file_path):
    target_keys = ["Sp", "Tp", "Yp", "acK", "Rm"]
    eps_factors = {key: [] for key in target_keys}
    s_factors = {key: [] for key in target_keys}

    workbook = openpyxl.load_workbook(file_path)
    
    eps_sheet = workbook["eps_factors"]
    s_sheet = workbook["s_factors"]
    
    #eps
    for column_index in range(1, eps_sheet.max_column + 1):
        first_cell_value = eps_sheet.cell(row=1, column=column_index).value
        for row_index in range(2,eps_sheet.max_row + 1):
            value_tmp = eps_sheet.cell(row=row_index, column=column_index).value
            eps_factors[first_cell_value].append(value_tmp)

    #s
    for column_index in range(1, s_sheet.max_column + 1):
        first_cell_value = s_sheet.cell(row=1, column=column_index).value
        for row_index in range(2,s_sheet.max_row + 1):
            value_tmp = s_sheet.cell(row=row_index, column=column_index).value
            s_factors[first_cell_value].append(value_tmp)

    return eps_factors, s_factors

def create_l(workbook,l_list,sheet_name="l"):
    sheet = workbook.create_sheet(sheet_name)
    sheet.append([""] + ref_sort)
    for i, row_label in enumerate(ref_sort):
        row_values = [row_label] 
        for j, _ in enumerate(ref_sort):
            val = (l_list[i] + l_list[j]) / 2
            row_values.append(val) 
        sheet.append(row_values)
    return workbook

def create_s(workbook,s_list,sheet_name="s"):
    sheet = workbook.create_sheet(sheet_name)
    sheet.append([""] + ref_sort)
    for i, row_label in enumerate(ref_sort):
        row_values = [row_label] 
        for j, _ in enumerate(ref_sort):
            val = (s_list[i] + s_list[j]) / 2
            row_values.append(val) 
        sheet.append(row_values)
    return workbook

residues_raw = [
        ["Sp", "SEP", 167.08, -2, 20],
        ["Tp", "TPO", 181.11, -2, 21],
        ["Yp", "PTR", 243.18, -2, 22],
        ["acK", "ALY", 170.17, 0, 23],
        ["Rm", "DA2", 184.19, 1, 24]
    ]

def add_residues(workbook,sheet_name="residues"):
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found.")
    
    sheet = workbook[sheet_name]
    start_row = sheet.max_row + 1
    for row_data in residues_raw:
            sheet.append(row_data)

    return workbook

def add_l(workbook,eps_factors,sheet_name="l"):
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found.")
    
    sheet = workbook[sheet_name]
    ref_keys = ["S", "T", "Y", "Q", "R"]
    ref_data = {key: [] for key in ref_keys}
    for column_index in range(1, 22):
        first_cell_value = sheet.cell(row=1, column=column_index).value
        if first_cell_value in ref_keys:
            for row_index in range(2,22):
                value_tmp = sheet.cell(row=row_index, column=column_index).value
                ref_data[first_cell_value].append(value_tmp)
    
    #Sp
    sheet.cell(row=1,column=22).value = "Sp"
    sheet.cell(row=22,column=1).value = "Sp"
    for i in range(0,20):
        value_tmp = 1 - (1 - ref_data["S"][i])**eps_factors["Sp"][i]
        sheet.cell(row=i+2,column=22).value = value_tmp
        sheet.cell(row=22,column=i+2).value = value_tmp
    #Sp-Sp, Tp-Sp, Yp-Sp, acK-Sp, Rm-Sp
    sheet.cell(row=22,column=22).value = 1 - (1 - ref_data["S"][13])**eps_factors["Sp"][20]
    sheet.cell(row=23,column=22).value = 1 - (1 - ref_data["S"][7])**eps_factors["Sp"][21]
    sheet.cell(row=24,column=22).value = 1 - (1 - ref_data["S"][18])**eps_factors["Sp"][22]
    sheet.cell(row=25,column=22).value = 1 - (1 - ref_data["S"][16])**eps_factors["Sp"][23]
    sheet.cell(row=26,column=22).value = 1 - (1 - ref_data["S"][2])**eps_factors["Sp"][24]

    #Tp
    sheet.cell(row=1,column=23).value = "Tp"
    sheet.cell(row=23,column=1).value = "Tp"
    for i in range(0,20):
        value_tmp = 1 - (1 - ref_data["T"][i])**eps_factors["Tp"][i]
        sheet.cell(row=i+2,column=23).value = value_tmp
        sheet.cell(row=23,column=i+2).value = value_tmp
    #Sp-Tp, Tp-Tp, Yp-Tp, acK-Tp, Rm-Tp
    sheet.cell(row=22,column=23).value = 1 - (1 - ref_data["T"][13])**eps_factors["Tp"][20]
    sheet.cell(row=23,column=23).value = 1 - (1 - ref_data["T"][7])**eps_factors["Tp"][21]
    sheet.cell(row=24,column=23).value = 1 - (1 - ref_data["T"][18])**eps_factors["Tp"][22]
    sheet.cell(row=25,column=23).value = 1 - (1 - ref_data["T"][16])**eps_factors["Tp"][23]
    sheet.cell(row=26,column=23).value = 1 - (1 - ref_data["T"][2])**eps_factors["Tp"][24]
    #Yp
    sheet.cell(row=1,column=24).value = "Yp"
    sheet.cell(row=24,column=1).value = "Yp"
    for i in range(0,20):
        value_tmp = 1 - (1 - ref_data["Y"][i])**eps_factors["Yp"][i]
        sheet.cell(row=i+2,column=24).value = value_tmp
        sheet.cell(row=24,column=i+2).value = value_tmp
    #Sp-Yp, Tp-Yp, Yp-Yp, acK-Yp, Rm-Yp
    sheet.cell(row=22,column=24).value = 1 - (1 - ref_data["Y"][13])**eps_factors["Yp"][20]
    sheet.cell(row=23,column=24).value = 1 - (1 - ref_data["Y"][7])**eps_factors["Yp"][21]
    sheet.cell(row=24,column=24).value = 1 - (1 - ref_data["Y"][18])**eps_factors["Yp"][22]
    sheet.cell(row=25,column=24).value = 1 - (1 - ref_data["Y"][16])**eps_factors["Yp"][23]
    sheet.cell(row=26,column=24).value = 1 - (1 - ref_data["Y"][2])**eps_factors["Yp"][24]
    #acK
    sheet.cell(row=1,column=25).value = "acK"
    sheet.cell(row=25,column=1).value = "acK"
    for i in range(0,20):
        value_tmp = 1 - (1 - ref_data["Q"][i])**eps_factors["acK"][i]
        sheet.cell(row=i+2,column=25).value = value_tmp
        sheet.cell(row=25,column=i+2).value = value_tmp
    #Sp-acK, Tp-acK, Yp-acK, acK-acK, Rm-acK
    sheet.cell(row=22,column=25).value = 1 - (1 - ref_data["Q"][13])**eps_factors["acK"][20]
    sheet.cell(row=23,column=25).value = 1 - (1 - ref_data["Q"][7])**eps_factors["acK"][21]
    sheet.cell(row=24,column=25).value = 1 - (1 - ref_data["Q"][18])**eps_factors["acK"][22]
    sheet.cell(row=25,column=25).value = 1 - (1 - ref_data["Q"][16])**eps_factors["acK"][23]
    sheet.cell(row=26,column=25).value = 1 - (1 - ref_data["Q"][2])**eps_factors["acK"][24]
    #Rm
    sheet.cell(row=1,column=26).value = "Rm"
    sheet.cell(row=26,column=1).value = "Rm"
    for i in range(0,20):
        value_tmp = 1 - (1 - ref_data["R"][i])**eps_factors["Rm"][i]
        sheet.cell(row=i+2,column=26).value = value_tmp
        sheet.cell(row=26,column=i+2).value = value_tmp
    #Sp-Rm, Tp-Rm, Yp-Rm, acK-Rm, Rm-Rm
    sheet.cell(row=22,column=26).value = 1 - (1 - ref_data["R"][13])**eps_factors["Rm"][20]
    sheet.cell(row=23,column=26).value = 1 - (1 - ref_data["R"][7])**eps_factors["Rm"][21]
    sheet.cell(row=24,column=26).value = 1 - (1 - ref_data["R"][18])**eps_factors["Rm"][22]
    sheet.cell(row=25,column=26).value = 1 - (1 - ref_data["R"][16])**eps_factors["Rm"][23]
    sheet.cell(row=26,column=26).value = 1 - (1 - ref_data["R"][2])**eps_factors["Rm"][24]
    
    return workbook

def add_s(workbook,s_factors,sheet_name="s"):
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found.")
    
    sheet = workbook[sheet_name]
    ref_keys = ["S", "T", "Y", "Q", "R"]
    ref_data = {key: [] for key in ref_keys}
    for column_index in range(1, 22):
        first_cell_value = sheet.cell(row=1, column=column_index).value
        if first_cell_value in ref_keys:
            for row_index in range(2,22):
                value_tmp = sheet.cell(row=row_index, column=column_index).value
                ref_data[first_cell_value].append(value_tmp)
    
    #Sp
    sheet.cell(row=1,column=22).value = "Sp"
    sheet.cell(row=22,column=1).value = "Sp"
    for i in range(0,20):
        value_tmp = ref_data["S"][i] + s_factors["Sp"][i]
        sheet.cell(row=i+2,column=22).value = value_tmp
        sheet.cell(row=22,column=i+2).value = value_tmp
    #Sp-Sp, Tp-Sp, Yp-Sp, acK-Sp, Rm-Sp
    sheet.cell(row=22,column=22).value = ref_data["S"][13] + s_factors["Sp"][20]
    sheet.cell(row=23,column=22).value = ref_data["S"][7] + s_factors["Sp"][21]
    sheet.cell(row=24,column=22).value = ref_data["S"][18] + s_factors["Sp"][22]
    sheet.cell(row=25,column=22).value = ref_data["S"][16] + s_factors["Sp"][23]
    sheet.cell(row=26,column=22).value = ref_data["S"][2] + s_factors["Sp"][24]
    #Tp
    sheet.cell(row=1,column=23).value = "Tp"
    sheet.cell(row=23,column=1).value = "Tp"
    for i in range(0,20):
        value_tmp = ref_data["T"][i] + s_factors["Tp"][i]
        sheet.cell(row=i+2,column=23).value = value_tmp
        sheet.cell(row=23,column=i+2).value = value_tmp
    #Sp-Tp, Tp-Tp, Yp-Tp, acK-Tp, Rm-Tp
    sheet.cell(row=22,column=23).value = ref_data["T"][13] + s_factors["Tp"][20]
    sheet.cell(row=23,column=23).value = ref_data["T"][7] + s_factors["Tp"][21]
    sheet.cell(row=24,column=23).value = ref_data["T"][18] + s_factors["Tp"][22]
    sheet.cell(row=25,column=23).value = ref_data["T"][16] + s_factors["Tp"][23]
    sheet.cell(row=26,column=23).value = ref_data["T"][2] + s_factors["Tp"][24]
    #Yp
    sheet.cell(row=1,column=24).value = "Yp"
    sheet.cell(row=24,column=1).value = "Yp"
    for i in range(0,20):
        value_tmp = ref_data["Y"][i] + s_factors["Yp"][i]
        sheet.cell(row=i+2,column=24).value = value_tmp
        sheet.cell(row=24,column=i+2).value = value_tmp
    #Sp-Yp, Tp-Yp, Yp-Yp, acK-Yp, Rm-Yp
    sheet.cell(row=22,column=24).value = ref_data["Y"][13] + s_factors["Yp"][20]
    sheet.cell(row=23,column=24).value = ref_data["Y"][7] + s_factors["Yp"][21]
    sheet.cell(row=24,column=24).value = ref_data["Y"][18] + s_factors["Yp"][22]
    sheet.cell(row=25,column=24).value = ref_data["Y"][16] + s_factors["Yp"][23]
    sheet.cell(row=26,column=24).value = ref_data["Y"][2] + s_factors["Yp"][24]
    #acK
    sheet.cell(row=1,column=25).value = "acK"
    sheet.cell(row=25,column=1).value = "acK"
    for i in range(0,20):
        value_tmp = ref_data["Q"][i] + s_factors["acK"][i]
        sheet.cell(row=i+2,column=25).value = value_tmp
        sheet.cell(row=25,column=i+2).value = value_tmp
    #Sp-acK, Tp-acK, Yp-acK, acK-acK, Rm-acK
    sheet.cell(row=22,column=25).value = ref_data["Q"][13] + s_factors["acK"][20]
    sheet.cell(row=23,column=25).value = ref_data["Q"][7] + s_factors["acK"][21]
    sheet.cell(row=24,column=25).value = ref_data["Q"][18] + s_factors["acK"][22]
    sheet.cell(row=25,column=25).value = ref_data["Q"][16] + s_factors["acK"][23]
    sheet.cell(row=26,column=25).value = ref_data["Q"][2] + s_factors["acK"][24]
    #Rm
    sheet.cell(row=1,column=26).value = "Rm"
    sheet.cell(row=26,column=1).value = "Rm"
    for i in range(0,20):
        value_tmp = ref_data["R"][i] + s_factors["Rm"][i]
        sheet.cell(row=i+2,column=26).value = value_tmp
        sheet.cell(row=26,column=i+2).value = value_tmp
    #Sp-Rm, Tp-Rm, Yp-Rm, acK-Rm, Rm-Rm
    sheet.cell(row=22,column=26).value = ref_data["R"][13] + s_factors["Rm"][20]
    sheet.cell(row=23,column=26).value = ref_data["R"][7] + s_factors["Rm"][21]
    sheet.cell(row=24,column=26).value = ref_data["R"][18] + s_factors["Rm"][22]
    sheet.cell(row=25,column=26).value = ref_data["R"][16] + s_factors["Rm"][23]
    sheet.cell(row=26,column=26).value = ref_data["R"][2] + s_factors["Rm"][24]
    
    return workbook

if __name__ == "__main__":
    import argparse
    import os
    parser = argparse.ArgumentParser(
        description="For HPS like force fields.\n Patch your own ptm parameter according to the given HPS model parameter file.\n \
        The example of the input parameter filecould be seen as residues.csv.\n \
        Please make sure the input file has the same format with residues.csv."
    )
    parser.add_argument(
        "--input_path", 
        type=str,
        help="The input HPS force field parameter file. (csv format)")
    parser.add_argument(
        "--output_path",
        type=str,
        help="The patched ptm HPS force field parameter file."
    )
    parser.add_argument(
        "--factor_path",
        type=str,
        default="factors.xlsx",
        help="The file that contains the scaling factors."
    )
    args = parser.parse_args()
    factor_path = args.factor_path
    input_path = args.input_path
    output_path = args.output_path
    # check
    if not os.path.isfile(factor_path):
        parser.error(f"ERROR: factor file does not exist: {factor_path}")
    if not os.path.isfile(input_path):
        parser.error(f"ERROR input file does not exist: {input_path}")

    eps_f, s_f = read_factors(factor_path)
    workbook, l_list, s_list = create_workbook(input_path)
    workbook = create_l(workbook, l_list)
    workbook = create_s(workbook, s_list)
    workbook = add_l(workbook, eps_f)
    workbook = add_s(workbook, s_f)
    workbook.save(output_path)
    print(f"Patched file: {output_path}")
